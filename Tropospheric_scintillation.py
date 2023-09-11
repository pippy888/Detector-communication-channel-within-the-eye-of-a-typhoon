import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
import math
import shutil

def Approximate_formula_for_saturated_water_vapor_pressure(temperature):
    return 6.11 * 10.0 ** ((7.5 * temperature) / (237.7 + temperature))

def getTableData(path):
    table_data = load_workbook(path)
    table_data_save = []
    sheet = table_data.active
    nrow = -1
    ncol = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # 遍历单元格
        for cell in row:
            # 打印单元格值
            if i == 0:
                ncol += 1
            elif i != 0:
                if type(cell) is str:
                    cell = list(cell)
                    cell[0] = '-'
                    cell = ''.join(cell)
                    cell = float(cell)
                table_data_save.append(cell)
                # print("{:<15}".format(cell), end='')
        nrow += 1
        # print()
    table_data_save = np.array(table_data_save)
    table_data_save = table_data_save.reshape(nrow, ncol)
    return table_data_save


def calculateS_for_waterVapour(b1, b2, e, p, q):
    return b1 * (10 ** -1) * e * q ** 3.5 * math.exp(b2 * (1 - q))

def calculate_width_line_for_waterVapour(b3, b4, b5, b6, e, p, q, f0):
    # 计算谱线宽度
    width_line = b3 * (10 ** -4) * (p * (q ** b4) + b5 * e * (q ** b6))
    width_line = 0.535 * width_line + math.sqrt(0.217 * width_line ** 2 + (2.1316 * (10 ** -12) * (f0 ** 2) / q))
    return width_line

def calculate_Correction_factor_for_waterVapour():
    return 0

def calculateF(e,p,q,f0,f,b3,b4,b5,b6):
    width_line = calculate_width_line_for_waterVapour(b3, b4, b5, b6, e, p, q, f0)
    Correction_factor = calculate_Correction_factor_for_waterVapour()
    # 正式计算谱线形状因子
    fi = (f / f0) * (((width_line - Correction_factor * (f0 - f)) / ((f0 - f) ** 2 + width_line ** 2)) + (
        (width_line - Correction_factor * (f0 + f)) / ((f0 + f) ** 2 + width_line ** 2)))
    return fi

def calculateNwet(e,p,q,table_data_waterVapour,f):
    sum_sy_for_waterVapour = 0
    for rows in table_data_waterVapour:
        f0 = rows[0]
        b1 = rows[1]
        b2 = rows[2]
        b3 = rows[3]
        b4 = rows[4]
        b5 = rows[5]
        b6 = rows[6]

        si_waterVapour = calculateS_for_waterVapour(b1, b2, e, p, q)
        fi_waterVapour = calculateF(e, p, q, f0, f , b3=b3, b4=b4, b5=b5, b6=b6)
        sum_sy_for_waterVapour += si_waterVapour * fi_waterVapour
    N_waterVapour = sum_sy_for_waterVapour
    return N_waterVapour

def calculate_standard_deviation_for_signal(Nwat):
    return 3.6 * 10 ** -3 + 10 ** -4 * Nwat

def calculate_L(hl,degree):
    x1 = 2 * hl
    x2 = math.sqrt(math.sin(degree) ** 2 + 2.35 * 10 ** -4) + math.sin(degree)
    return x1 / x2
def calculate_x(Deff,f,L):
    x = 1.22 * Deff ** 2 * (f / L)
    return x
def calculate_g(Deff,f,L):
    x = calculate_x(Deff,f,L)
    g = 3.86 * (x ** 2 + 1) ** (11 / 12) * math.sin(11 / 6 * math.atan(1 / x)) - 7.08 * x ** (5 / 6)
    return math.sqrt(g)

def calculate_standard_deviation_for_path(standard_deviation_for_signal,f,g,degree):
    return standard_deviation_for_signal * f ** (7 / 12) * g / abs((math.sin(degree)) ** 1.2)

def calculate_ap(p):
    x1 = -0.061 * (math.log10(p)) ** 3
    x2 = 0.072 * (math.log10(p)) ** 2
    x3 = -1.71 * math.log10(p)
    ap = x1 + x2 + x3 + 3.0
    return ap

def calculateKw():
    pl = 10
    # 福州维度 26 C lat = 0 Co = 76
    Clat = 0
    Co = 76
    Kw = pl ** 1.5 * 10 ** ((Clat + Co) / 10)
    return Kw

def calculateA_242(Kw, f, degree, percentage_p):
    x1 = 10 * math.log10(Kw)
    x2 = 9 * math.log10(f)
    x3 = 55 * math.log10(1 + degree) # 单位是mrad
    x4 = 10 * math.log(percentage_p)
    A_242 = x1 + x2 - x3 - x4
    return A_242
def calculate_elevation(A,Kw,f,p):
    x1 = Kw * f ** 0.9
    x2 = p * 10 ** (A / 10)
    return (x1 / x2) ** (1 / 5.5) - 1

def calculate_deputy_A(e,elevation):
    return -1 * (55 / (1 + elevation)) * math.log10(e)

def calculate_deputy_A2(A2,x,degree,Deff,f,hl):
    degree = degree * 10 ** 3 # mrad
    x1 = calculate_gx_compare_deputy_gx(x)
    x2 = calculate_dx_compare_ddegree(Deff, f, hl, degree)
    return A2 * (x1 * x2 - (1.2 / math.tan(degree))) * (1 / 1000)
def calculate_gx_compare_deputy_gx(x):
    x1 = (11 / 6) * math.atan(1 / x)
    x2 = 1770 * (x ** 2 + 1) + 2123 * x ** (1 / 6) * (x ** 2 + 1) ** (11 / 12) * (math.cos(x1) - x * math.sin(x1))
    x3 = 12 * x ** (1 / 6) * (x ** 2 + 1) * (354 * x ** (5 / 6) - 193 * (x ** 2 + 1) ** (11 / 12) * math.sin(x1))
    return x2 / x3
def calculate_dx_compare_ddegree(Deff, f, hl, degree):
    x1 = 1.22 * Deff ** 2 * f / (2 * hl)
    x2 = math.sin(degree) / math.sqrt((math.sin(degree)) ** 2 + 2.35 * 10 ** -4)
    return x1 * (x2 + 1) * math.cos(degree)

def calculate_Flickering_fading(A1, deputy_A1, A2, deputy_A2, elevation1, elevation2, degree):
    elevation_difference = elevation2 - elevation1
    ap = deputy_A1 / A1
    bp = (math.log(A2 / A1) - ap * elevation_difference) / (elevation_difference ** 2)
    yp = (deputy_A2 - A2 * (ap + 2 * bp * elevation_difference)) / (A2 * elevation_difference ** 2)
    x1 = ap * (degree - elevation1)
    x2 = bp * (degree - elevation1) ** 2
    x3 = yp * (degree - elevation1) ** 2 * (degree - elevation2)
    Ap = A1 * math.exp(x1 + x2 + x3)
    return Ap

if __name__ == '__main__':
    heigth_interval = [0, 150, 1000, 3000, 5000, 7500, 10000, 12000, 14000, 16000, 18000, 20000, 22000]
    heigth_interval = [x * 0.001 for x in heigth_interval]  # 公里为单位
    # 采用58847(福州）研究所在2019年7月测得的平均数据
    # 总气压
    pressure = np.array([995., 932.5, 829.33333333, 641., 493., 350., 250., 200., 150., 100.5, 70., 50., 25., ])
    # 温度
    temperature = np.array([29., 24., 18., 8., -5.5, -21.5, -39., -50., -66., -76.5, -68., -63., -56.])
    # 得出水汽分压e
    e = Approximate_formula_for_saturated_water_vapor_pressure(temperature)
    # 总气压减去水汽分压可以得到干空气气压p
    p = pressure - e
    # 参数 q = 300K / temperature
    temperature_K = temperature + 273.15
    q = np.full((len(heigth_interval)), 300) / temperature_K
    f = 40
    hl = 1000
    percentage_p = 0.01
    degree = math.degrees(4)
    table_data_oxygen = getTableData('.\\氧气衰减的谱线数据.xlsx')
    table_data_waterVapour = getTableData('.\\水汽衰减的谱线数据.xlsx')

    # 计算2.4.1
    # 计算N（waterVapour）使用0km的数据计算 未处理的无线电波折射率
    N_wat = calculateNwet(e[0],p[0],q[0],table_data_waterVapour,f)
    # 计算信号赋值的标准偏差
    standard_deviation_for_signal = calculate_standard_deviation_for_signal(N_wat)
    # 1000m 以km为单位 角度设置成5 计算有效路径长度L
    L = calculate_L(hl,degree)
    # 有效天线直径
    Deff = 0.1
    # 计算天线平均系数g
    g = calculate_g(Deff,f,L)
    # 计算适用期间内和路径上信号的标准偏差
    standard_deviation_for_path = calculate_standard_deviation_for_path(standard_deviation_for_signal,f,g,degree)
    # 计算时间百分比p = 0.01时间百分比系数ap
    ap = calculate_ap(percentage_p)
    # 计算超出时间百分比p%内的衰减深度
    A_241 = ap * standard_deviation_for_path
    print(A_241)

    # 计算2.4.2
    # 计算年份平均最坏月份的地面系数Kw
    Kw = calculateKw()
    A_242 = calculateA_242(Kw,f,degree,percentage_p)
    print(A_242)
    # 计算2.4.3
    # 仰角小于5时闪烁/多路径衰减中弱衰减部分分布的计算
    # set A1 = 25dB
    A1 = 25
    elevation1 = calculate_elevation(A1, Kw, f, percentage_p)
    deputy_A1 = calculate_deputy_A(e[0],elevation1)
    A2 = A_241
    x = calculate_x(Deff,f,L)
    deputy_A2 = calculate_deputy_A2(A2, x, degree, Deff, f, hl)
    elevation2 = calculate_elevation(A2, Kw, f, percentage_p)
    A_243 = calculate_Flickering_fading(A1, deputy_A1, A2, deputy_A2, elevation1, elevation2, degree)
    print(A_243)