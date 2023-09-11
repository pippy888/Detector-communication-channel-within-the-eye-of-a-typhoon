import numpy as np
import os
import shutil
import docx
from openpyxl import load_workbook
import math


# pressure是总的气压，已经得到，根据温度可以算出水汽分压e，二者相减，可以得到干空气气压p
# 根据温度得到水汽分压的公式如下：注意温度此时是摄氏度为单位
def Approximate_formula_for_saturated_water_vapor_pressure(temperature):
    return 6.11 * 10.0 ** ((7.5 * temperature) / (237.7 + temperature))


def getTableData(path):
    table_data = load_workbook(path)
    table_data_save = []
    sheet = table_data.active
    nrow = -1
    ncol = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # 遍历单元格
        for cell in row:
            # 打印单元格值
            if i == 0:
                ncol += 1
            elif i != 0:
                if type(cell) is str:
                    cell = list(cell)
                    cell[0] = '-'
                    cell = ''.join(cell)
                    cell = float(cell)
                table_data_save.append(cell)
                # print("{:<15}".format(cell), end='')
        nrow += 1
        # print()
    table_data_save = np.array(table_data_save)
    table_data_save = table_data_save.reshape(nrow, ncol)
    return table_data_save


def calculateY(e, p, q, table_data_oxygen, table_data_waterVapour, f):
    # 先计算累和
    sum_sy_for_oxygen = 0
    sun_sy_for_waterVapour = 0
    for rows in table_data_oxygen:  # 计算每一行
        f0 = rows[0]
        a1 = rows[1]
        a2 = rows[2]
        a3 = rows[3]
        a4 = rows[4]
        a5 = rows[5]
        a6 = rows[6]
        si_oxygen = calculateS_for_oxygen(a1, a2, e, p, q)
        fi_oxygen = calculateF('oxygen', f, f0, e, p, q, a3=a3, a4=a4, a5=a5, a6=a6)
        sum_sy_for_oxygen += si_oxygen * fi_oxygen
    # N_oxygen后面的参数是由气压造成的氮吸收以及德拜频谱产生的干空气连续吸收谱
    NDF = calculateNDF(f, e, p, q)

    N_oxygen = sum_sy_for_oxygen + NDF

    for rows in table_data_waterVapour:
        f0 = rows[0]
        b1 = rows[1]
        b2 = rows[2]
        b3 = rows[3]
        b4 = rows[4]
        b5 = rows[5]
        b6 = rows[6]

        si_waterVapour = calculateS_for_waterVapour(b1, b2, e, p, q)
        fi_waterVapour = calculateF('waterVapour', f, f0, e, p, q, b3=b3, b4=b4, b5=b5, b6=b6)
        sun_sy_for_waterVapour += si_waterVapour * fi_waterVapour
    N_waterVapour = sun_sy_for_waterVapour

    Y = 0.1820 * f * (N_oxygen + N_waterVapour)
    return Y


def calculateS_for_oxygen(a1, a2, e, p, q):
    return a1 * (10 ** -7) * p * (q ** 3) * math.exp(a2 * (1 - q))


def calculateS_for_waterVapour(b1, b2, e, p, q):
    return b1 * (10 ** -1) * e * q ** 3.5 * math.exp(b2 * (1 - q))


def calculate_width_line_for_oxygen(a3, a4, e, p, q):
    # 计算谱线宽度
    width_line = a3 * (10 ** -4) * (p * q ** (0.8 - a4) + 1.1 * e * q)
    width_line = math.sqrt(width_line ** 2 + 2.25 * 10 ** -6)
    return width_line


def calculate_width_line_for_waterVapour(b3, b4, b5, b6, e, p, q, f0):
    # 计算谱线宽度
    width_line = b3 * (10 ** -4) * (p * (q ** b4) + b5 * e * (q ** b6))
    width_line = 0.535 * width_line + math.sqrt(0.217 * width_line ** 2 + (2.1316 * (10 ** -12) * (f0 ** 2) / q))
    return width_line


def calculate_Correction_factor_for_oxygen(a5, a6, e, p, q):
    Correction_factor = (a5 + a6 * q) * 10 ** -4 * (p + e) * q ** 0.8
    return Correction_factor


def calculate_Correction_factor_for_waterVapour():
    return 0


def calculateF(classString, f, f0, e, p, q, a3=0, a4=0, a5=0, a6=0, b3=0, b4=0, b5=0, b6=0):
    if classString == 'oxygen':
        # 计算谱线宽度
        width_line = calculate_width_line_for_oxygen(a3, a4, e, p, q)
        # 计算修正因子
        Correction_factor = calculate_Correction_factor_for_oxygen(a5, a6, e, p, q)
    else:
        width_line = calculate_width_line_for_waterVapour(b3, b4, b5, b6, e, p, q, f0)
        Correction_factor = calculate_Correction_factor_for_waterVapour()
    # 正式计算谱线形状因子
    fi = (f / f0) * (((width_line - Correction_factor * (f0 - f)) / ((f0 - f) ** 2 + width_line ** 2)) + (
            (width_line - Correction_factor * (f0 + f)) / ((f0 + f) ** 2 + width_line ** 2)))
    return fi


def calculateNDF(f, e, p, q):
    # 先算出d
    d = 5.6 * (10 ** -4) * (p + e) * q ** 0.8

    x1 = f * p * q ** 2
    x2 = 6.14 * (10 ** -5) / (d * (1 + (f / d) ** 2))
    x3 = (1.4 * (10 ** -12) * p * (q ** 1.5)) / (1 + 1.9 * 10 ** -5 * f ** 1.5)
    NDF = x1 * (x2 + x3)
    return NDF


def calculateN(e, p, t):
    x1 = 77.6 * p / t
    x2 = 72 * e / t
    x3 = 3.75 * 10 ** 5 * e / t ** 2
    N = x1 + x2 + x3
    return 1 + N * 10 ** -6


def calculateB(n1, r1, ri, B1, e, p, t):
    ni = calculateN(e, p, t)
    Bi = math.asin(n1 * r1 * math.sin(B1) / (ni * ri))
    return Bi

def calculateA(ri,Bi,layer):
    x1 = -1 * ri * math.cos(Bi)
    x2 = ri ** 2 * math.cos(Bi) ** 2 + 2 * ri * layer + layer ** 2
    return x1 + math.sqrt(x2)

if __name__ == '__main__':
    heigth_interval = [0, 150, 1000, 3000, 5000, 7500, 10000, 12000, 14000, 16000, 18000, 20000, 22000]
    heigth_interval = [x * 0.001 for x in heigth_interval]  # 公里为单位
    # 采用58847(福州）研究所在2019年7月测得的平均数据
    # 总气压
    pressure = np.array([995., 932.5, 829.33333333, 641., 493., 350., 250., 200., 150., 100.5, 70., 50., 25., ])
    # 温度
    temperature = np.array([29., 24., 18., 8., -5.5, -21.5, -39., -50., -66., -76.5, -68., -63., -56.])
    # 得出水汽分压e
    e = Approximate_formula_for_saturated_water_vapor_pressure(temperature)
    # 总气压减去水汽分压可以得到干空气气压p
    p = pressure - e
    # 参数 q = 300K / temperature
    temperature_K = temperature + 273.15
    q = np.full((len(heigth_interval)), 300) / temperature_K
    f = 40
    # f为指定值 是40GHz

    table_data_oxygen = getTableData('.\\氧气衰减的谱线数据.xlsx')
    table_data_waterVapour = getTableData('.\\水汽衰减的谱线数据.xlsx')

    # 对于每个高度，都要算出对应的气体衰减y
    heigth_interval_y = []
    for i in range(len(heigth_interval)):
        Y = calculateY(e[i], p[i], q[i], table_data_oxygen, table_data_waterVapour, f)
        heigth_interval_y.append(Y)

    # 至此已经计算出所有的y
    # 下面开始计算路径衰减

    # tan值是高度（22km）除以距离（50km）
    # 全部使用弧度，而不是角度
    tan_degrees = 22 / 50
    tan_degrees = math.atan(tan_degrees)
    B1 = math.degrees(90) - tan_degrees
    n1 = calculateN(e[0], p[0], temperature_K[0])
    r1 = 6371
    # 之所以减1，是因为按厚度计算衰减，高度分层下，厚度比高度分层节点数少一个，衰减使用起点节点
    sum_ay = 0
    for i in range(len(heigth_interval) - 1):
        layer = heigth_interval[i + 1] - heigth_interval[i]
        ri = r1 + heigth_interval[i]
        Bi = calculateB(n1,r1,ri,B1,e[i],p[i],temperature_K[i])
        ai = calculateA(ri,Bi,layer)
        sum_ay += ai * heigth_interval_y[i]

    with open('.\\gas_attenuaion.txt', 'w') as f:
        for i in heigth_interval_y:
            f.write(str(i) + '\n')
        f.write(str(sum_ay))